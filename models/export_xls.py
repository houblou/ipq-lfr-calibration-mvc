
"""

Structure Excel:
  Colonne A    : index i (1-30) + étiquettes de synthèse
  Colonnes B+  : une colonne par série (les 2 premières = acquisition initiale COM1/COM2)
  Lignes 2-31 : N[1] à N[30]
  Ligne 32     : MOYENNE
  Ligne 33     : VARIANCE
  Ligne 34     : T moy (°C)
  Ligne 35     : HR moy (%)
  Ligne 36     : Distance (mm)
  Ligne 37     : Date
  Ligne 38     : Heure début
"""
import os
from datetime import datetime
from typing import List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from core.logger import creer_logger
from core.paths import nom_fichier_xls
from core.config import NB_POINTS

logger = creer_logger("phase3")

# Le bloc de données occupe les lignes 2..(nb_points+1). Les lignes de synthèse
# suivent (offset après nb_points). Pour 30 points : MOYENNE en 32, etc. Le tout
# est calculé dynamiquement par ExportXLS (overlock — nombre de points variable).
ROW_DATA_START = 2
_OFFSETS_SYNTH = [   # (offset après nb_points, label)
    (2, "MEAN"), (3, "VARIANCE"), (4, "Mean T (°C)"), (5, "Mean RH (%)"),
    (6, "Distance (mm)"), (7, "Date"), (8, "Start time"), (9, "Operator"),
]
# Constantes par défaut (30 points) — conservées pour référence externe / tests.
ROW_MOYENNE    = NB_POINTS + 2
ROW_VARIANCE   = NB_POINTS + 3
ROW_T_MOY      = NB_POINTS + 4
ROW_HR_MOY     = NB_POINTS + 5
ROW_DISTANCE   = NB_POINTS + 6
ROW_DATE       = NB_POINTS + 7
ROW_HEURE      = NB_POINTS + 8
ROW_OPERATEUR  = NB_POINTS + 9


class ExportXLS:
    """Gère la création et l'écriture du fichier Excel multi-séries."""

    def __init__(self, indice_notation: str, dossier: str = ".",
                 simulation: bool = False, operateur: str = "",
                 nb_points: int = NB_POINTS) -> None:
        self.indice_notation = indice_notation
        self.dossier         = dossier
        self.chemin_fichier  = os.path.join(dossier, nom_fichier_xls(indice_notation))
        self.wb              = None
        self.ws              = None
        self.serie_courante  = 0
        self.simulation      = simulation
        self.operateur       = operateur
        # Mise en page calculée d'après le nombre de points (overlock).
        self.nb_points       = int(nb_points)
        self._labels_synth   = {self.nb_points + off: lbl for off, lbl in _OFFSETS_SYNTH}
        self.row_moyenne     = self.nb_points + 2
        self.row_variance    = self.nb_points + 3
        self.row_t_moy       = self.nb_points + 4
        self.row_hr_moy      = self.nb_points + 5
        self.row_distance    = self.nb_points + 6
        self.row_date        = self.nb_points + 7
        self.row_heure       = self.nb_points + 8
        self.row_operateur   = self.nb_points + 9

    def ouvrir(self) -> bool:
        """Crée le classeur Excel et prépare la feuille principale."""
        if not OPENPYXL_OK:
            logger.error("openpyxl non installé — export XLS impossible.")
            return False
        try:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "SIMULATION" if self.simulation else "Measurements"
            self._ecrire_entete()
            if self.simulation:
                warning = self.ws.cell(
                    row=self.row_operateur + 1, column=1,
                    value="SIMULATION DATA — NOT VALID FOR METROLOGY",
                )
                warning.font = Font(bold=True, color="FFFFFF")
                warning.fill = PatternFill("solid", fgColor="C00000")
            os.makedirs(self.dossier, exist_ok=True)
            self.wb.save(self.chemin_fichier)
            logger.info("Fichier Excel créé : %s", self.chemin_fichier)
            return True
        except Exception as exc:
            logger.error("Erreur création Excel : %s", exc)
            return False

    def ajouter_serie(
        self,
        n:        List[Optional[float]],
        m:        float,
        v:        float,
        t_moy:    float = 0.0,
        hr_moy:   float = 0.0,
        distance: float = 0.0,
        date:     str   = "",
        heure:    str   = "",
        label:    Optional[str] = None,
    ) -> None:
        """
        Ajoute une série dans la colonne suivante.
        Sauvegarde automatiquement après chaque série.
        """
        if self.ws is None:
            raise RuntimeError("Excel file is not open.")

        if len(n) != self.nb_points:
            raise ValueError(
                f"A series must contain exactly {self.nb_points} points; received: {len(n)}.")
        # Les points invalides (None) sont CONSERVES et marques 'INVALID' (jamais
        # supprimes) ; ils sont deja exclus du calcul M/V en amont.
        perdus = sum(1 for valeur in n if valeur is None)

        if not date:
            date = datetime.now().strftime("%d/%m/%Y")
        if not heure:
            heure = datetime.now().strftime("%H:%M:%S")

        self.serie_courante += 1
        col = self.serie_courante + 1   # B = col 2, C = col 3…

        # En-tête de colonne (label explicite ou numérotation automatique)
        entete = label if label else f"Series {self.serie_courante}"
        if perdus:
            entete += f"  ({perdus} invalid)"
        cell_h = self.ws.cell(row=1, column=col, value=entete)
        cell_h.font      = Font(bold=True, color="FFFFFF")
        cell_h.fill      = PatternFill("solid", fgColor="2E75B6")
        cell_h.alignment = Alignment(horizontal="center")

        # Données N[i] (lignes 2 à 31) — point perdu ecrit 'INVALID' + surligne
        for i, valeur in enumerate(n, start=1):
            if valeur is None:
                cell = self.ws.cell(row=i + 1, column=col, value="INVALID")
                cell.fill = PatternFill("solid", fgColor="F8CBAD")
            else:
                self.ws.cell(row=i + 1, column=col, value=valeur)

        # Synthèse
        synthese = {
            self.row_moyenne:  m,
            self.row_variance: v,
            self.row_t_moy:    t_moy,
            self.row_hr_moy:   hr_moy,
            self.row_distance: distance,
            self.row_date:     date,
            self.row_heure:    heure,
        }
        for row, valeur in synthese.items():
            cell = self.ws.cell(row=row, column=col, value=valeur)
            cell.fill = PatternFill("solid", fgColor="FFF3CD")

        self._sauvegarder()
        logger.info(
            "%s ajoutée (col %d) — M=%.6f V=%.6f T_moy=%.2f HR_moy=%.2f dist=%.1f",
            entete, col, m, v, t_moy, hr_moy, distance,
        )

    def fermer(self) -> None:
        """Sauvegarde finale."""
        self._sauvegarder()
        logger.info("Fichier Excel finalisé : %s", self.chemin_fichier)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _ecrire_entete(self) -> None:
        """Écrit la colonne A : index i + étiquettes de synthèse."""
        bold = Font(bold=True)

        # Ligne 1, col A
        self.ws.cell(row=1, column=1, value="i").font = bold

        # Index 1..nb_points
        for i in range(1, self.nb_points + 1):
            self.ws.cell(row=i + 1, column=1, value=i)

        # Étiquettes de synthèse
        fill_synth = PatternFill("solid", fgColor="D9E1F2")
        for row, label in self._labels_synth.items():
            cell = self.ws.cell(row=row, column=1, value=label)
            cell.font = bold
            cell.fill = fill_synth

        # Opérateur — donnée de session, colonne B fixe (une seule valeur)
        if self.operateur:
            cell_op = self.ws.cell(row=self.row_operateur, column=2, value=self.operateur)
            cell_op.fill = PatternFill("solid", fgColor="FFF3CD")

        # Largeur colonne A
        self.ws.column_dimensions["A"].width = 16

    def _sauvegarder(self) -> None:
        if self.wb and self.chemin_fichier:
            try:
                self.wb.save(self.chemin_fichier)
            except Exception as exc:
                logger.exception("Excel save failed: %s", exc)
                raise RuntimeError(
                    f"Unable to save the Excel file: {self.chemin_fichier}"
                ) from exc
