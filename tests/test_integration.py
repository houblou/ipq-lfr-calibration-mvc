# -*- coding: utf-8 -*-
"""
test_integration.py — 10 tests d'intégration :
ports/communication, interface, arborescence, lancement des fonctions, exports.
"""
import glob
import math
import os
import tempfile
import time
import unittest

from openpyxl import load_workbook

from models.ports import GestionPorts
from models import stats
from models.thermo import ThermoService
from models.export_xls import ExportXLS, ROW_MOYENNE, ROW_OPERATEUR

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _attendre(cond, timeout=2.0):
    t0 = time.time()
    while time.time() - t0 < timeout:
        if cond():
            return True
        time.sleep(0.02)
    return False


class _FakeGP:
    def __init__(self, simulation=True, connecte=False):
        self.mode_simulation = simulation
        self.thermo1 = type("P", (), {"is_open": True})() if connecte else None

    def thermo_actif(self):
        return self.mode_simulation or self.thermo1 is not None

    def lire_thermo(self):
        return (22.5, 48.0)


# ── 1. Arborescence : toute la structure s'importe ─────────────────────────────
class TestArborescence(unittest.TestCase):
    def test_packages_et_classes_importables(self):
        import core.config, core.logger, core.paths  # noqa: F401
        from models.acquisition import Acquisition
        from models.calibration import BoucleCalibration, GestionAudio
        from models.initialisation import GestionInitialisation
        from models.audit import JournalAudit
        from models.detection import detecter
        from controllers.init_controller import InitController
        from controllers.mesure_controller import MesureController
        from controllers.admin_controller import AdminController
        from controllers.connexion_controller import ConnexionController
        from controllers.thermo_controller import ThermoController
        from views.app_window import ApplicationIPQ
        from views.monitor import MonitorTab
        for obj in (Acquisition, BoucleCalibration, GestionAudio, GestionInitialisation,
                    JournalAudit, detecter, InitController, MesureController, AdminController,
                    ConnexionController, ThermoController, ApplicationIPQ, MonitorTab):
            self.assertTrue(callable(obj))


# ── 2 & 3. Invariants MVC vérifiés sur le disque ───────────────────────────────
class TestInvariantsMVC(unittest.TestCase):
    def test_modeles_et_core_sans_tkinter(self):
        for dossier in ("models", "core"):
            for f in glob.glob(os.path.join(RACINE, dossier, "*.py")):
                with open(f, encoding="utf-8") as fh:
                    src = fh.read()
                for interdit in ("import tkinter", "from tkinter", "import ttk",
                                 "import matplotlib", "from matplotlib"):
                    self.assertNotIn(interdit, src,
                                     f"{os.path.basename(f)} importe '{interdit}'")

    def test_controllers_sans_widgets(self):
        for f in glob.glob(os.path.join(RACINE, "controllers", "*.py")):
            with open(f, encoding="utf-8") as fh:
                src = fh.read()
            for interdit in ("import tkinter", "from tkinter", ".pack(", ".grid(",
                             "tk.Label", "tk.Button"):
                self.assertNotIn(interdit, src,
                                 f"{os.path.basename(f)} contient '{interdit}'")


# ── 4 & 5. Ports / communication ───────────────────────────────────────────────
class TestPorts(unittest.TestCase):
    def test_lecture_simulation(self):
        gp = GestionPorts()
        self.assertIsNone(gp.lire_com("com1"))
        with self.assertRaises(RuntimeError):
            gp.lire_thermo()
        gp.set_simulation_mode(True)
        self.assertIsInstance(gp.lire_com("com1"), float)
        t, hr = gp.lire_thermo()
        self.assertIsInstance(t, float)
        self.assertIsInstance(hr, float)
        gp.set_simulation_mode(False)
        self.assertIsNone(gp.lire_com("com1"))

    def test_connexion_port_invalide(self):
        gp = GestionPorts()
        self.assertFalse(gp.connecter_serie("COM_INEXISTANT_999", "com1"))
        self.assertTrue(gp.get_erreurs())


# ── 6. Stats (lancement + parité + bords) ──────────────────────────────────────
class TestStats(unittest.TestCase):
    def test_fonctions_et_bords(self):
        self.assertEqual(stats.moyenne_variance([1.0, 2.0, 3.0, 4.0]), (2.5, 1.25))
        self.assertEqual(stats.moyenne([1.0, None, 3.0]), 2.0)
        self.assertEqual(stats.moyenne_variance([]), (0.0, 0.0))
        self.assertIsNone(stats.moyenne([None, None]))
        m, sigma = stats.moyenne_sigma([1.0, 2.0, 3.0, 4.0])
        self.assertAlmostEqual(m, 2.5)
        self.assertAlmostEqual(sigma, math.sqrt(1.25))
        self.assertTrue(stats.alerte_variance(0.02, 0.01))
        self.assertFalse(stats.alerte_variance(None, 0.01))


# ── 7. Service thermo (se lance, pousse les bons callbacks) ─────────────────────
class TestServiceThermo(unittest.TestCase):
    def test_simulation_indispo_suspendu(self):
        r = []
        s = ThermoService(_FakeGP(simulation=True), est_occupe=lambda: False, intervalle_s=0.03)
        s.demarrer(lambda t, hr: r.append((t, hr)), lambda: r.append("indispo"),
                   lambda e: r.append("err"))
        self.assertTrue(_attendre(lambda: any(isinstance(x, tuple) for x in r)))
        s.arreter()

        r2 = []
        s2 = ThermoService(_FakeGP(simulation=False, connecte=False),
                           est_occupe=lambda: False, intervalle_s=0.03)
        s2.demarrer(lambda t, hr: r2.append("m"), lambda: r2.append("indispo"),
                    lambda e: r2.append("err"))
        self.assertTrue(_attendre(lambda: "indispo" in r2))
        s2.arreter()

        r3 = []
        s3 = ThermoService(_FakeGP(simulation=True), est_occupe=lambda: True, intervalle_s=0.03)
        s3.demarrer(lambda t, hr: r3.append("m"), lambda: r3.append("i"),
                    lambda e: r3.append("e"))
        time.sleep(0.12)
        s3.arreter()
        self.assertEqual(r3, [])


# ── 8 & 9. Exports Excel ───────────────────────────────────────────────────────
class TestExports(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.tmp.cleanup()

    def test_deux_series(self):
        export = ExportXLS("TEST", dossier=self.tmp.name, operateur="Operator")
        self.assertTrue(export.ouvrir())
        pts1 = [float(i) for i in range(1, 31)]
        pts2 = [float(i) * 2 for i in range(1, 31)]
        export.ajouter_serie(pts1, 15.5, 74.9, 20.1, 49.8, 500.0, "22/06/2026", "10:00:00", label="S1")
        export.ajouter_serie(pts2, 31.0, 299.7, 20.2, 49.5, 500.0, "22/06/2026", "10:30:00", label="S2")
        self.assertEqual(export.serie_courante, 2)
        ws = load_workbook(export.chemin_fichier, data_only=True).active
        self.assertEqual(ws.cell(row=1, column=2).value, "S1")
        self.assertEqual(ws.cell(row=1, column=3).value, "S2")
        self.assertEqual(ws.cell(row=ROW_MOYENNE, column=2).value, 15.5)
        self.assertEqual(ws.cell(row=ROW_MOYENNE, column=3).value, 31.0)
        self.assertEqual(ws.cell(row=31, column=3).value, 60.0)

    def test_simulation_filigrane_et_operateur(self):
        export = ExportXLS("TEST", dossier=self.tmp.name, simulation=True, operateur="Op")
        self.assertTrue(export.ouvrir())
        ws = load_workbook(export.chemin_fichier).active
        self.assertEqual(ws.title, "SIMULATION")
        self.assertEqual(ws.cell(row=ROW_OPERATEUR, column=2).value, "Op")


# ── 10. Interface : la Vue expose tout ce que les contrôleurs appellent ────────
class TestInterfaceVue(unittest.TestCase):
    METHODES_REQUISES = [
        "afficher_avertissement", "afficher_erreur", "demander_secret", "demander_confirmation",
        "_lire_nb_series", "_lire_attente_s", "_get_distance", "_verifier_ports_requis",
        "_maj_init_pt", "_vue_init_demarrage", "_vue_init_standby", "_vue_init_interrompu",
        "_vue_init_resultats", "_vue_init_pret", "_vue_init_approuve",
        "_vue_mesure_demarrage", "_vue_mesure_attente", "_vue_mesure_serie",
        "_vue_mesure_arret_demande", "_vue_mesure_boutons_repos", "_vue_mesure_erreur",
        "_vue_mesure_interrompu", "_vue_mesure_termine",
        "_vue_admin_deverrouille", "_vue_admin_verrouille", "_vue_simulation",
        "_vue_cle_configuree", "_vue_audit_resultat", "_actualiser_btn_sim",
        "_vue_ports_disponibles", "_vue_port_ok", "_vue_port_echec",
        "_vue_detect_indispo", "_vue_detect_scan", "_vue_detection",
        "_vue_operateur", "_vue_connexion_ok",
        "_log", "_badge", "_statut", "_naviguer", "after",
    ]

    def test_methodes_attendues_par_les_controleurs_existent(self):
        import tkinter as tk
        from views.app_window import ApplicationIPQ
        self.assertTrue(issubclass(ApplicationIPQ, tk.Tk))
        manquants = [m for m in self.METHODES_REQUISES if not hasattr(ApplicationIPQ, m)]
        self.assertEqual(manquants, [], f"Méthodes manquantes sur la Vue : {manquants}")


if __name__ == "__main__":
    unittest.main()
