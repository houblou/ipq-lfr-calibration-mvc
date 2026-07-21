# -*- coding: utf-8 -*-
import os
import tempfile
import unittest
from unittest import mock

from openpyxl import load_workbook

from models.acquisition import Acquisition
from models.ports import GestionPorts
from models.export_xls import (
    ExportXLS,
    ROW_DATE,
    ROW_DISTANCE,
    ROW_HEURE,
    ROW_HR_MOY,
    ROW_MOYENNE,
    ROW_T_MOY,
    ROW_VARIANCE,
)
from models.calibration import BoucleCalibration
from models import security


class TestCalculsMetrologiques(unittest.TestCase):
    def test_moyenne_et_variance_population(self):
        moyenne, variance = Acquisition._calculer_stats([1.0, 2.0, 3.0, 4.0])
        self.assertAlmostEqual(moyenne, 2.5)
        self.assertAlmostEqual(variance, 1.25)

    def test_calcul_ignore_les_lectures_invalides(self):
        moyenne, variance = Acquisition._calculer_stats([1.0, None, 3.0])
        self.assertAlmostEqual(moyenne, 2.0)
        self.assertAlmostEqual(variance, 1.0)


class TestSimulationSecurity(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.config_file = os.path.join(self.tempdir.name, "security.json")
        self.patch_dir = mock.patch.object(security, "CONFIG_DIR", self.tempdir.name)
        self.patch_file = mock.patch.object(security, "CONFIG_FILE", self.config_file)
        self.patch_dir.start()
        self.patch_file.start()

    def tearDown(self):
        self.patch_file.stop()
        self.patch_dir.stop()
        self.tempdir.cleanup()

    def test_admin_key_is_hashed_and_verified(self):
        security.configure_admin_key("a-secure-admin-key")
        self.assertTrue(security.admin_key_configured())
        self.assertTrue(security.verify_admin_key("a-secure-admin-key"))
        self.assertFalse(security.verify_admin_key("wrong-key"))
        with open(self.config_file, "r", encoding="utf-8") as stream:
            self.assertNotIn("a-secure-admin-key", stream.read())

    def test_rejects_short_admin_key(self):
        with self.assertRaises(ValueError):
            security.configure_admin_key("short")

    def test_synthetic_values_require_explicit_simulation_mode(self):
        ports = GestionPorts()
        self.assertIsNone(ports.lire_com("com1"))
        with self.assertRaises(RuntimeError):
            ports.lire_thermo()

        ports.set_simulation_mode(True)
        self.assertIsInstance(ports.lire_com("com1"), float)
        temperature, humidity = ports.lire_thermo()
        self.assertIsInstance(temperature, float)
        self.assertIsInstance(humidity, float)

        ports.set_simulation_mode(False)
        self.assertIsNone(ports.lire_com("com1"))


class TestExportExcel(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.export = ExportXLS("TEST", dossier=self.tempdir.name)
        self.assertTrue(self.export.ouvrir())

    def tearDown(self):
        self.tempdir.cleanup()

    def test_preserve_les_30_points_et_la_synthese(self):
        points = [float(i) for i in range(1, 31)]
        self.export.ajouter_serie(
            points, 15.5, 74.9166666667, 20.1, 49.8, 500.0,
            "22/06/2026", "10:15:00", label="Init COM1",
        )

        feuille = load_workbook(self.export.chemin_fichier, data_only=True).active
        self.assertEqual(feuille.cell(row=31, column=2).value, 30.0)
        self.assertEqual(feuille.cell(row=ROW_MOYENNE, column=2).value, 15.5)
        self.assertAlmostEqual(feuille.cell(row=ROW_VARIANCE, column=2).value, 74.9166666667)
        self.assertEqual(feuille.cell(row=ROW_T_MOY, column=2).value, 20.1)
        self.assertEqual(feuille.cell(row=ROW_HR_MOY, column=2).value, 49.8)
        self.assertEqual(feuille.cell(row=ROW_DISTANCE, column=2).value, 500.0)
        self.assertEqual(feuille.cell(row=ROW_DATE, column=2).value, "22/06/2026")
        self.assertEqual(feuille.cell(row=ROW_HEURE, column=2).value, "10:15:00")

    def test_refuse_une_serie_de_mauvaise_longueur(self):
        # nb_points = CAPACITE de la feuille : on rejette le DEPASSEMENT et la serie
        # VIDE. Une serie plus COURTE que la capacite est valide (une X-serie de 20
        # points dans une feuille de 30 : init/finale restent a 30).
        with self.assertRaises(ValueError):
            self.export.ajouter_serie([1.0] * (self.export.nb_points + 1), 1.0, 0.0)  # depassement
        with self.assertRaises(ValueError):
            self.export.ajouter_serie([], 1.0, 0.0)                                   # vide

    def test_point_invalide_conserve_et_marque(self):
        # Un point None est CONSERVE et marque 'INVALID' (jamais supprime),
        # et l'en-tete de colonne signale le nombre d'invalides.
        self.export.ajouter_serie([1.0] * 29 + [None], 1.0, 0.0, label="S")
        feuille = load_workbook(self.export.chemin_fichier, data_only=True).active
        self.assertEqual(feuille.cell(row=31, column=2).value, "INVALID")
        self.assertIn("invalid", str(feuille.cell(row=1, column=2).value).lower())

    def test_ouvrir_dimensionne_a_la_taille_demandee(self):
        # Création à l'approbation : tant que ouvrir() n'est pas appelé, le fichier n'existe
        # pas (est_ouvert False). ouvrir(nb_points=...) fixe la CAPACITÉ à la création.
        export = ExportXLS("SIZE", dossier=self.tempdir.name)
        self.assertFalse(export.est_ouvert())
        self.assertTrue(export.ouvrir(nb_points=40))
        self.assertTrue(export.est_ouvert())
        self.assertEqual(export.nb_points, 40)
        self.assertEqual(export.row_moyenne, 42)
        export.ajouter_serie([float(i) for i in range(1, 41)], 5.0, 1.0, label="S1")
        feuille = load_workbook(export.chemin_fichier, data_only=True).active
        self.assertEqual(feuille.cell(row=41, column=2).value, 40.0)     # 40e point
        self.assertEqual(feuille.cell(row=42, column=2).value, 5.0)      # MEAN aligné en 42

    def test_ouvrir_echec_laisse_non_ouvert_et_retry_propre(self):
        # Atomicité : si la sauvegarde échoue (fichier verrouillé), ouvrir() renvoie False
        # ET est_ouvert() reste False -> un retry recrée un classeur PROPRE (serie_courante
        # remis à 0), évitant que la ré-écriture des colonnes d'init soit sautée.
        import openpyxl
        export = ExportXLS("ATOM", dossier=self.tempdir.name)
        with mock.patch.object(openpyxl.Workbook, "save", side_effect=OSError("locked")):
            self.assertFalse(export.ouvrir(nb_points=30))
        self.assertFalse(export.est_ouvert())
        self.assertTrue(export.ouvrir(nb_points=30))    # retry sans verrou
        self.assertTrue(export.est_ouvert())
        self.assertEqual(export.serie_courante, 0)

    def test_overlock_layout_dynamique(self):
        # Overlock : un nombre de points != 30 décale la synthèse en conséquence.
        export = ExportXLS("OVL", dossier=self.tempdir.name, nb_points=10)
        self.assertTrue(export.ouvrir())
        export.ajouter_serie([float(i) for i in range(1, 11)], 5.5, 8.25, label="S")
        feuille = load_workbook(export.chemin_fichier, data_only=True).active
        self.assertEqual(export.row_moyenne, 12)                       # 10 pts -> MOYENNE en 12
        self.assertEqual(feuille.cell(row=11, column=2).value, 10.0)   # dernier point en 11
        self.assertEqual(feuille.cell(row=export.row_moyenne, column=2).value, 5.5)


class _AcquisitionFactice:
    def __init__(self):
        self.arrete = False

    def demander_arret(self):
        self.arrete = True

    def lancer(self, cible="com1", callback_point=None, callback_fin=None):
        return [1.0] * 30, 1.0, 0.0, 20.0, 50.0


class TestBoucleCalibration(unittest.TestCase):
    def test_signale_une_erreur_export_et_appelle_la_fin(self):
        fins = []

        def export_en_echec(*args, **kwargs):
            raise RuntimeError("fichier verrouillé")

        boucle = BoucleCalibration(
            nb_series=2,
            acquisition=_AcquisitionFactice(),
            fn_export=export_en_echec,
            callback_fin=lambda resultats: fins.append(resultats),
        )
        boucle._boucle()

        self.assertEqual(boucle.resultats, [])
        self.assertEqual(fins, [[]])
        self.assertIn("fichier verrouillé", boucle.erreur)


if __name__ == "__main__":
    unittest.main()
